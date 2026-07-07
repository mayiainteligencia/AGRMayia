"""
===============================================================================
perfilar.py  -  Perfilador de datos para los Excel del cliente (AgroMayia)
===============================================================================

Autor : Martin Cortes
Rol   : Ingenieria de datos - AgroMayia (AgroBotanical ERP)

Lo que hace este script:
  Le paso un archivo Excel por la terminal y me devuelve un diagnostico de
  cada hoja: filas, columnas, tipos, vacios (NULL), posibles llaves, datos
  personales (PII, como correos) e inconsistencias de mayusculas. Y lo mas
  importante para mi: me dice que hojas son fuente de datos (tablas limpias)
  y cuales son pivotes o reportes que NO debo importar.

Mis dos reglas de oro aqui:
  1. No relleno vacios ni invento datos. Solo observo y reporto. El vacio
     (NULL) es informacion: me dice donde el cliente tiene huecos.
  2. Un detector automatico solo senala; la decision final la tomo yo.

Como lo uso:
  python perfilar.py mi-archivo.xlsx
  python perfilar.py uploads/REG-FENOLOGICO_2026.xlsx

  El diagnostico se imprime en pantalla y ademas se guarda en la carpeta
  reportes/.
===============================================================================
"""

import sys
import os
from datetime import datetime
import warnings

import pandas as pd
import openpyxl

# Silencio los avisos de openpyxl sobre extensiones raras de Excel; no me afectan.
warnings.filterwarnings("ignore")


# -----------------------------------------------------------------------------
# Datos de autoria. Los imprimo en el reporte para dejar rastro de quien y
# cuando lo genero.
# -----------------------------------------------------------------------------
AUTOR = "Martin Cortes"

# -----------------------------------------------------------------------------
# Ajustes. Estos son mis "umbrales": las reglas que yo decido. Si algun dia
# quiero afinar el detector, cambio estos numeros aqui y no toco el resto.
# -----------------------------------------------------------------------------
COLS_PARA_SOSPECHAR_PIVOTE = 40   # una hoja con muchisimas columnas me huele a reporte
MAX_FILAS_MUESTRA = None          # None = leo todo; pongo un numero para probar rapido


def separador(caracter="=", n=74):
    """Me dibuja una linea para separar secciones dentro del reporte."""
    return caracter * n


def detectar_pii(nombre_columna, serie):
    """
    Reviso si una columna trae datos personales (PII).
    Por ahora busco correos, que caen bajo la LFPDPPP y no deben ir a Git.
    Devuelvo True si mas de la mitad de los valores parecen correo.
    """
    muestra = serie.dropna().astype(str).head(50)
    if len(muestra) == 0:
        return False
    parecen_correo = muestra.str.contains("@", regex=False).sum()
    return parecen_correo > len(muestra) / 2


def detectar_casing_inconsistente(serie):
    """
    Cazo el problema de 'tiffany' vs 'Tiffany': valores que son el mismo texto
    salvo por mayusculas/minusculas o espacios. Devuelvo la lista de grupos en
    conflicto (por ejemplo ['a@x.com', 'A@x.com']).
    """
    texto = serie.dropna().astype(str)
    if len(texto) == 0:
        return []
    # Agrupo por su version normalizada (minusculas y sin espacios de sobra).
    normalizado = texto.str.strip().str.lower()
    conflictos = []
    for clave, grupo in texto.groupby(normalizado):
        originales = grupo.unique()
        if len(originales) > 1:  # mismo valor normalizado, pero escrito de varias formas
            conflictos.append(list(originales))
    return conflictos


def clasificar_hoja(ws, df):
    """
    El corazon del script: decido si una hoja es FUENTE DE DATOS o PIVOTE/REPORTE.

    Un pivote o reporte no lo importo: es un dibujo de datos que en realidad
    viven en otra hoja. Me apoyo en varias "senales". Entre mas senales junte,
    mas seguro estoy. Devuelvo (veredicto, lista_de_razones).
    """
    razones = []

    # Senal 1: la hoja tiene graficas dibujadas adentro.
    n_graficas = len(getattr(ws, "_charts", []))
    if n_graficas > 0:
        razones.append(f"contiene {n_graficas} grafica(s) incrustada(s)")

    # Senal 2: tiene celdas combinadas (merged). Una tabla limpia casi nunca las
    # usa; en cambio un reporte con titulos y encabezados bonitos, si.
    n_merged = len(ws.merged_cells.ranges)
    if n_merged > 3:
        razones.append(f"{n_merged} celdas combinadas (layout de reporte)")

    # Senal 3: es rarisimamente ancha (muchas columnas). Los pivotes crecen a lo
    # ancho: una columna por semana, por mes, etc.
    if df.shape[1] >= COLS_PARA_SOSPECHAR_PIVOTE:
        razones.append(f"{df.shape[1]} columnas (demasiado ancha, tipico de pivote)")

    # Senal 4: muchos encabezados sin nombre ('Unnamed') que ademas estan casi
    # vacios. Ojo con esto: una tabla limpia puede traer una mini-tabla-leyenda
    # pegada al lado (columnas sin nombre pero casi vacias), y eso NO la vuelve
    # pivote. Por eso solo cuento las columnas sin nombre que esten vacias >90%:
    # esas si me delatan que la fila 1 no es un encabezado real.
    sin_nombre_vacias = sum(
        1 for c in df.columns
        if str(c).startswith("Unnamed")
        and (df[c].isnull().mean() > 0.9 if len(df) else True)
    )
    if df.shape[1] > 0 and sin_nombre_vacias > df.shape[1] / 2:
        razones.append(f"{sin_nombre_vacias} encabezados sin nombre y vacios (fila 1 no es tabla)")

    # Senal 5: la hoja esta casi vacia (poca sustancia).
    if df.shape[0] < 3:
        razones.append("muy pocas filas de datos")

    # Mi decision: exijo 2 o mas senales para declarar 'pivote'. Con una sola
    # senal puedo equivocarme (por ejemplo, la tabla-leyenda pegada al lado). Los
    # pivotes de verdad casi siempre disparan varias senales a la vez. Cuando hay
    # una sola senal, lo dejo como fuente pero marcado para que yo lo revise.
    if len(razones) >= 2:
        veredicto = "PIVOTE / REPORTE (no importar)"
    elif len(razones) == 1:
        veredicto = "FUENTE DE DATOS (revisar - 1 senal dudosa)"
    else:
        veredicto = "FUENTE DE DATOS (importable)"
    return veredicto, razones


def perfilar_hoja(ws, df, nombre):
    """Armo el bloque de diagnostico de UNA hoja."""
    lineas = []
    lineas.append(separador("-"))
    lineas.append(f"HOJA: {nombre}")

    veredicto, razones = clasificar_hoja(ws, df)
    lineas.append(f"  VEREDICTO: {veredicto}")
    if razones:
        for r in razones:
            lineas.append(f"     - {r}")

    filas, cols = df.shape
    lineas.append(f"  Dimensiones: {filas} filas x {cols} columnas")

    # Si es pivote, no me detengo a perfilar columna por columna: paro aqui.
    if veredicto.startswith("PIVOTE"):
        lineas.append("  (Omito el detalle de columnas: esta hoja no es fuente.)")
        return "\n".join(lineas)

    # Detalle columna por columna (solo para las fuentes de datos).
    lineas.append("  Columnas:")
    lineas.append(f"     {'#':>2}  {'NOMBRE':<24} {'TIPO':<12} {'VACIOS':>8}  NOTAS")
    for i, col in enumerate(df.columns, start=1):
        serie = df[col]
        tipo = str(serie.dtype)
        n_vacios = serie.isnull().sum()
        pct_vacios = (n_vacios / filas * 100) if filas else 0
        vacios_txt = f"{n_vacios} ({pct_vacios:.0f}%)"

        notas = []
        # Cuento cuantos valores distintos hay; me sirve para detectar posibles llaves.
        n_unicos = serie.nunique(dropna=True)
        if n_unicos == filas and filas > 0:
            notas.append("posible LLAVE (todo unico)")
        # Reviso PII.
        if detectar_pii(col, serie):
            notas.append("PII (correos) - NO subir a Git")
        # Reviso casing inconsistente.
        conflictos = detectar_casing_inconsistente(serie)
        if conflictos:
            ejemplo = conflictos[0][:2]
            notas.append(f"casing inconsistente ej. {ejemplo}")

        nota_txt = "; ".join(notas) if notas else ""
        nombre_corto = str(col)[:24]
        lineas.append(f"     {i:>2}  {nombre_corto:<24} {tipo:<12} {vacios_txt:>8}  {nota_txt}")

    # Resumo las columnas con demasiados vacios para preguntarlas al cliente.
    altos = [(c, df[c].isnull().sum() / filas * 100) for c in df.columns if filas]
    altos = [(c, p) for c, p in altos if p >= 30]
    if altos:
        lineas.append("  Aviso - columnas con MUCHOS vacios (>=30%), preguntar al cliente:")
        for c, p in altos:
            lineas.append(f"     - {c}: {p:.0f}% vacio")

    return "\n".join(lineas)


def perfilar_archivo(ruta):
    """Mi funcion principal: abro el Excel y perfilo todas sus hojas."""
    if not os.path.exists(ruta):
        print(f"ERROR: no encuentro el archivo '{ruta}'")
        sys.exit(1)

    nombre_archivo = os.path.basename(ruta)
    fecha_ejecucion = datetime.now().strftime("%Y-%m-%d %H:%M")

    salida = []
    salida.append(separador())
    salida.append(f"PERFILADO DE DATOS - {nombre_archivo}")
    salida.append(f"Hecho por : {AUTOR}")
    salida.append(f"Ejecutado : {fecha_ejecucion}")
    salida.append(separador())

    # Abro con openpyxl para poder ver graficas y celdas combinadas.
    wb = openpyxl.load_workbook(ruta, data_only=True)
    # Y con pandas para analizar el contenido tabular.
    hojas_pandas = pd.read_excel(ruta, sheet_name=None, nrows=MAX_FILAS_MUESTRA)

    salida.append(f"\nEl archivo tiene {len(wb.sheetnames)} hoja(s): {wb.sheetnames}\n")

    fuentes, pivotes = [], []
    for nombre in wb.sheetnames:
        ws = wb[nombre]
        df = hojas_pandas.get(nombre, pd.DataFrame())
        bloque = perfilar_hoja(ws, df, nombre)
        salida.append(bloque)
        salida.append("")  # linea en blanco de separacion

        veredicto, _ = clasificar_hoja(ws, df)
        (pivotes if veredicto.startswith("PIVOTE") else fuentes).append(nombre)

    # Resumen final: mi chuleta para decidir que importar.
    salida.append(separador())
    salida.append("RESUMEN - que hago con cada hoja")
    salida.append(separador())
    salida.append(f"  [Fuentes de datos] importar {len(fuentes)}: {fuentes}")
    salida.append(f"  [Pivotes/reportes] ignorar {len(pivotes)}: {pivotes}")
    salida.append("")
    salida.append("  Recordatorio: los vacios se quedan como NULL. Nunca los relleno")
    salida.append("  con '-' en el dato. El '-' es solo para el frontend, al final.")
    salida.append("")
    salida.append(separador("-"))
    salida.append(f"Reporte generado por {AUTOR} el {fecha_ejecucion}.")

    reporte = "\n".join(salida)

    # Lo imprimo en pantalla.
    print(reporte)

    # Y lo guardo en la carpeta reportes/.
    carpeta_reportes = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reportes")
    os.makedirs(carpeta_reportes, exist_ok=True)
    nombre_reporte = f"perfil_{os.path.splitext(nombre_archivo)[0]}.txt"
    ruta_reporte = os.path.join(carpeta_reportes, nombre_reporte)
    with open(ruta_reporte, "w", encoding="utf-8") as f:
        f.write(reporte)
    print(f"\n>>> Reporte guardado en: reportes/{nombre_reporte}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("USO: python perfilar.py <archivo.xlsx>")
        print("Ejemplo: python perfilar.py uploads/REG-FENOLOGICO_2026.xlsx")
        sys.exit(1)
    perfilar_archivo(sys.argv[1])